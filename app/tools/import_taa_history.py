"""Einmaliger Import der Momentum-Historie aus TAA Conviction.xlsx.

Liest den ``Momentum``-Sheet, extrahiert pro Ticker die woechentlichen
Klassifikationen (Golden Cross / Kurs > SMA-200 / Kurs < SMA-200 / Death
Cross) und persistiert sie per UPSERT in ``sector_momentum_snapshots``.

Usage:
    python -m app.tools.import_taa_history "TAA Conviction.xlsx"
    python -m app.tools.import_taa_history "TAA Conviction.xlsx" --dry-run
    python -m app.tools.import_taa_history "TAA Conviction.xlsx" --sheet Momentum
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime

import openpyxl
import pandas as pd

from app.core.momentum import MOMENTUM_STATES
from app.core.persistence import save_sector_snapshot
from app.core.sectors import (
    GROUP_INDUSTRY,
    GROUP_SECTOR,
    display_name,
    group_for,
)


# Layout der beiden parallelen Tabellen im Sheet "Momentum".
# Row 4 enthaelt die Header (Ticker, Gruppe, 5 Datumsspalten).
# Sektoren stehen in Spalten B..H, Industrien in Spalten J..P.
_HEADER_ROW = 4

_SECTOR_LAYOUT = {
    "group": GROUP_SECTOR,
    "ticker_col": 2,  # B
    "date_cols": (4, 5, 6, 7, 8),  # D..H
    "row_range": range(5, 16),  # 11 Sektor-ETFs
}

_INDUSTRY_LAYOUT = {
    "group": GROUP_INDUSTRY,
    "ticker_col": 10,  # J
    "date_cols": (12, 13, 14, 15, 16),  # L..P
    "row_range": range(5, 24),  # 19 Industrie-ETFs
}


def _parse_date_headers(ws, date_cols: tuple[int, ...]) -> list[date]:
    out: list[date] = []
    for col in date_cols:
        value = ws.cell(row=_HEADER_ROW, column=col).value
        if isinstance(value, datetime):
            out.append(value.date())
        elif isinstance(value, date):
            out.append(value)
        else:
            raise ValueError(
                f"Unerwarteter Datumswert in Header-Zelle "
                f"{ws.cell(row=_HEADER_ROW, column=col).coordinate}: {value!r}"
            )
    return out


def _collect_rows(ws, layout: dict) -> list[dict]:
    dates = _parse_date_headers(ws, layout["date_cols"])
    rows: list[dict] = []
    for r in layout["row_range"]:
        raw_ticker = ws.cell(row=r, column=layout["ticker_col"]).value
        if not raw_ticker:
            continue
        ticker = str(raw_ticker).strip()
        grp = group_for(ticker)
        if grp is None or grp != layout["group"]:
            print(
                f"  [warn] Zeile {r}: Ticker '{ticker}' nicht in der "
                f"Gruppe {layout['group']!r} erwartet; übersprungen.",
                file=sys.stderr,
            )
            continue
        name = display_name(ticker)
        for col, d in zip(layout["date_cols"], dates):
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            val = str(val).strip()
            if val not in MOMENTUM_STATES:
                print(
                    f"  [warn] {ticker} @ {d:%Y-%m-%d}: unbekanntes Label "
                    f"'{val}', übersprungen.",
                    file=sys.stderr,
                )
                continue
            rows.append(
                {
                    "snapshot_date": d,
                    "ticker": ticker,
                    "group": grp,
                    "display_name": name,
                    "last_price": None,
                    "sma_50": None,
                    "sma_200": None,
                    "momentum": val,
                }
            )
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", help="Pfad zur TAA Conviction.xlsx")
    parser.add_argument(
        "--sheet",
        default="Momentum",
        help="Name des Sheets (Default: Momentum)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Nur anzeigen, nichts in die Datenbank schreiben.",
    )
    args = parser.parse_args(argv)

    wb = openpyxl.load_workbook(args.path, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(
            f"Fehler: Sheet '{args.sheet}' nicht gefunden. "
            f"Verfügbar: {wb.sheetnames}",
            file=sys.stderr,
        )
        return 2
    ws = wb[args.sheet]

    rows = _collect_rows(ws, _SECTOR_LAYOUT) + _collect_rows(ws, _INDUSTRY_LAYOUT)
    if not rows:
        print("Keine importierbaren Zeilen gefunden.", file=sys.stderr)
        return 1

    df = pd.DataFrame(rows)
    n_weeks = df["snapshot_date"].nunique()
    n_tickers = df["ticker"].nunique()
    print(
        f"Gefunden: {len(df)} Zeilen · {n_weeks} Wochen · {n_tickers} Ticker."
    )
    print("Verteilung je Woche:")
    for d, sub in df.groupby("snapshot_date"):
        counts = sub["momentum"].value_counts().to_dict()
        summary = " · ".join(f"{k}:{v}" for k, v in sorted(counts.items()))
        print(f"  {d:%Y-%m-%d} ({len(sub):3d})  {summary}")

    if args.dry_run:
        print("\nDry-Run: keine Schreiboperationen.")
        return 0

    total = 0
    for d, sub in df.groupby("snapshot_date"):
        frame = sub.drop(columns=["snapshot_date"]).reset_index(drop=True)
        n = save_sector_snapshot(frame, d)
        print(f"  -> {d:%Y-%m-%d}: {n} Zeilen gespeichert")
        total += n
    print(f"\nFertig. Insgesamt {total} Snapshots (UPSERT) persistiert.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
